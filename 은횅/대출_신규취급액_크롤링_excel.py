import asyncio
from playwright.async_api import async_playwright, Page
from bs4 import BeautifulSoup
import pandas as pd
import logging
import re
import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 로깅 설정
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# 수정된 헤더 (평균금리 제거)
DESIRED_HEADERS = [
    '대출 종류', '은행', '구분',
    '1000~951점', '950~901점', '900~851점', '850~801점',
    '800~751점', '750~701점', '700~651점', '650~601점', '600점이하',
    '평균신용점수', 'CB회사명', '참고사항'
]

async def crawl_kfb_loan_rates():
    async with async_playwright() as p:
        # 브라우저를 헤드리스 모드로 실행
        browser = await p.chromium.launch(headless=True)
        # 다운로드를 허용하지 않는 컨텍스트 생성
        context = await browser.new_context()
        page = await context.new_page()

        url = 'https://portal.kfb.or.kr/compare/loan_household_new.php'
        logging.info(f"페이지 접속: {url}")
        await page.goto(url, timeout=1000000)

        # 페이지가 완전히 로드될 때까지 대기
        await page.wait_for_load_state('networkidle')

        # 1. 모든 은행 선택 체크박스 클릭 (id="BankAll")
        try:
            await page.check('input#BankAll')
            logging.info("모든 은행 선택 체크박스 클릭 완료")
        except Exception as e:
            logging.error(f"체크박스 클릭 실패: {e}")

        # 2. 공시기준 선택 (신규취급액 기준: id="select_new_balance_1")
        try:
            await page.check('input#select_new_balance_1')
            logging.info("공시기준 '신규취급액 기준' 선택 완료")
        except Exception as e:
            logging.error(f"공시기준 선택 실패: {e}")

        # 3. 상세구분 선택 (대출금리 상세보기: id="all_show_1")
        try:
            await page.check('input#all_show_1')
            logging.info("상세구분 '대출금리 상세보기' 선택 완료")
        except Exception as e:
            logging.error(f"상세구분 선택 실패: {e}")

        # 4. 대출종류 선택
        loan_types = {
            '분할상환방식 주택담보대출': '#opt_1_1',
            '일시상환방식 주택담보대출': '#opt_1_2',
            '일반신용대출': '#opt_1_3',
            '신용한도대출(마이너스대출)': '#opt_1_4',
            '전세자금대출': '#opt_1_5',
            '가계대출': '#opt_1_6'
        }

        # 데이터 수집 리스트 초기화
        collected_data = []

        # 각 대출종류를 순회하며 데이터 수집
        for loan_label, loan_selector in loan_types.items():
            try:
                # 새로운 대출종류 선택
                await page.evaluate(f"document.querySelector('{loan_selector}').checked = true;")
                logging.info(f"대출종류 '{loan_label}' 선택 완료")
            except Exception as e:
                logging.error(f"대출종류 '{loan_label}' 선택 실패: {e}")
                continue  # 다음 대출종류로 넘어감

            # 5. 검색 버튼 클릭 (실제 클릭 이벤트 사용)
            try:
                await page.click('div.btnArea .btn a[href^="Javascript:LoanHouseholdNewSearch_cur();"]')
                logging.info("검색 버튼 클릭 완료")
                await asyncio.sleep(2)
            except Exception as e:
                logging.error(f"검색 버튼 클릭 실패: {e}")
                continue  # 다음 대출종류로 넘어감

            # 검색 결과 로드 대기 (결과 테이블이 업데이트될 때까지 대기)
            try:
                await page.wait_for_selector('table.resultList_ty02 tbody tr', timeout=60000)
                logging.info(f"검색 결과 테이블 로드 완료 for 대출종류 '{loan_label}'")
            except Exception as e:
                logging.error(f"검색 결과 테이블 로드 실패 for 대출종류 '{loan_label}': {e}")
                continue  # 다음 대출종류로 넘어감

            # 6. 페이지 콘텐츠 가져오기
            page_content = await page.content()
            soup = BeautifulSoup(page_content, 'html.parser')

            # 7. 결과 테이블 찾기
            table = soup.find('table', class_='resultList_ty02')
            if not table:
                logging.error(f"결과 테이블을 찾을 수 없습니다 for 대출종류 '{loan_label}'")
                continue  # 다음 대출종류로 넘어감

            # 8. 테이블 헤더 추출
            header_rows = table.find_all('tr')
            if len(header_rows) < 2:
                logging.error(f"헤더 행을 찾을 수 없습니다 for 대출종류 '{loan_label}'")
                continue

            # 두 번째 tr은 신용점수 구간 헤더
            second_header = header_rows[1]
            score_columns = [th.get_text(separator=' ', strip=True).replace('<br>', ' ').replace('\n', '').strip()
                             for th in second_header.find_all('th')]
            logging.info(f"신용점수 구간: {score_columns}")

            # 데이터 행 (세 번째 tr부터)
            data_rows = table.find_all('tr')[2:]

            # 은행 정보 및 추가 정보 유지 변수
            current_bank = None
            average_credit_score = ''
            cb_company = ''
            remarks = ''

            # 임시 변수로 데이터 수집
            temp_data = {}

            for tr in data_rows:
                tds = tr.find_all('td')
                if not tds:
                    continue  # 빈 행 스킵

                # 첫 번째 td에 은행 이름이 있는지 확인 (링크나 이미지 포함 여부)
                first_td = tds[0]
                if first_td.find('a') or first_td.find('img'):
                    # 새로운 은행 시작
                    current_bank = first_td.get_text(strip=True).replace('\xa0', ' ').strip()

                    # 구분: 두 번째 td
                    rate_type = tds[1].get_text(strip=True)

                    # CB사 신용점수별 금리 추출 (다음 9개 td)
                    cb_rates = [td.get_text(strip=True) for td in tds[2:11]]

                    # 추가 정보: 평균신용점수, CB회사명, 참고사항
                    average_credit_score = tds[11].get_text(strip=True) if len(tds) > 11 else ''
                    cb_company = tds[12].get_text(strip=True) if len(tds) > 12 else ''
                    remarks = tds[13].get_text(strip=True) if len(tds) > 13 else ''

                    # 임시 데이터 초기화
                    temp_data = {
                        '대출 종류': loan_label,
                        '은행': current_bank,
                        '구분': rate_type,
                        '1000~951점': cb_rates[0] if len(cb_rates) > 0 else '',
                        '950~901점': cb_rates[1] if len(cb_rates) > 1 else '',
                        '900~851점': cb_rates[2] if len(cb_rates) > 2 else '',
                        '850~801점': cb_rates[3] if len(cb_rates) > 3 else '',
                        '800~751점': cb_rates[4] if len(cb_rates) > 4 else '',
                        '750~701점': cb_rates[5] if len(cb_rates) > 5 else '',
                        '700~651점': cb_rates[6] if len(cb_rates) > 6 else '',
                        '650~601점': cb_rates[7] if len(cb_rates) > 7 else '',
                        '600점이하': cb_rates[8] if len(cb_rates) > 8 else '',
                        '평균신용점수': average_credit_score,
                        'CB회사명': cb_company,
                        '참고사항': remarks
                    }
                    collected_data.append(temp_data.copy())
                else:
                    # 이전 은행 정보 유지, 구분 및 금리만 업데이트
                    rate_type = tds[0].get_text(strip=True)
                    cb_rates = [td.get_text(strip=True) for td in tds[1:10]]

                    # 기존 데이터 업데이트
                    if temp_data:
                        temp_data['구분'] = rate_type
                        temp_data['1000~951점'] = cb_rates[0] if len(cb_rates) > 0 else ''
                        temp_data['950~901점'] = cb_rates[1] if len(cb_rates) > 1 else ''
                        temp_data['900~851점'] = cb_rates[2] if len(cb_rates) > 2 else ''
                        temp_data['850~801점'] = cb_rates[3] if len(cb_rates) > 3 else ''
                        temp_data['800~751점'] = cb_rates[4] if len(cb_rates) > 4 else ''
                        temp_data['750~701점'] = cb_rates[5] if len(cb_rates) > 5 else ''
                        temp_data['700~651점'] = cb_rates[6] if len(cb_rates) > 6 else ''
                        temp_data['650~601점'] = cb_rates[7] if len(cb_rates) > 7 else ''
                        temp_data['600점이하'] = cb_rates[8] if len(cb_rates) > 8 else ''

                        # 현재 은행 데이터 갱신
                        collected_data.append(temp_data.copy())

            logging.info(f"대출종류 '{loan_label}'의 데이터 수집 완료: {len(collected_data)} rows")

            # 다음 대출종류를 선택하기 전에 현재 선택된 대출종류를 해제
            try:
                await page.evaluate(f"document.querySelector('{loan_selector}').checked = false;")
                logging.info(f"대출종류 '{loan_label}' 해제 완료")
            except Exception as e:
                logging.error(f"대출종류 '{loan_label}' 해제 실패: {e}")

        # 브라우저 닫기
        await browser.close()

        df = pd.DataFrame(collected_data)
        df = df[DESIRED_HEADERS]

        # 12. 최종 엑셀 파일 저장
        try:
            now = datetime.datetime.now()
            timestamp = now.strftime("%Y%m%d_%H%M%S")
            file_name = f"대출금리_신규취급액기준_{timestamp}.xlsx"
            file_path = os.path.join(os.getcwd(), file_name)

            # 엑셀 파일 작성
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 데이터 작성 (헤더는 직접 작성할 것이므로 DataFrame만 작성)
                df.to_excel(writer, index=False, header=False, startrow=2, sheet_name='종합대출금리')

                # 워크시트 포맷팅
                workbook = writer.book
                worksheet = writer.sheets['종합대출금리']

                # 헤더 병합 및 설정
                # '대출 종류' (A1:A2)
                worksheet.merge_cells('A1:A2')
                worksheet.cell(row=1, column=1).value = '대출 종류'
                worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

                # '은행' (B1:B2)
                worksheet.merge_cells('B1:B2')
                worksheet.cell(row=1, column=2).value = '은행'
                worksheet.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')

                # '구분' (C1:C2)
                worksheet.merge_cells('C1:C2')
                worksheet.cell(row=1, column=3).value = '구분'
                worksheet.cell(row=1, column=3).alignment = Alignment(horizontal='center', vertical='center')

                # '신용점수별 금리(%)' (D1:M1)
                score_col_start = 4  # 'D' 열부터 시작
                num_score_columns = 10  # '1000~951점'부터 '600점이하'까지 10개
                score_col_end = score_col_start + num_score_columns - 1  # 'M' 열
                worksheet.cell(row=1, column=score_col_start).value = '신용점수별 금리(%)'
                worksheet.merge_cells(start_row=1, start_column=score_col_start, end_row=1, end_column=score_col_end)
                worksheet.cell(row=1, column=score_col_start).alignment = Alignment(horizontal='center', vertical='center')

                # 신용점수 구간 헤더 설정 (두 번째 행)
                score_headers = DESIRED_HEADERS[3:13]  # '1000~951점' ~ '600점이하'
                for idx, score in enumerate(score_headers):
                    col_idx = score_col_start + idx
                    cell = worksheet.cell(row=2, column=col_idx)
                    cell.value = score
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                # '평균신용점수', 'CB회사명', '참고사항' (N1:P1)
                additional_headers = DESIRED_HEADERS[12:]  # '평균신용점수', 'CB회사명', '참고사항'
                additional_col_start = score_col_end + 1  # 'N' 열부터 시작
                for idx, header in enumerate(additional_headers):
                    col_idx = additional_col_start + idx
                    worksheet.cell(row=1, column=col_idx).value = header
                    worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
                    worksheet.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

                # 각 열의 최대 길이를 기반으로 열 너비 조정
                total_columns = additional_col_start + len(additional_headers) - 1
                for col in range(1, total_columns + 1):
                    column_letter = get_column_letter(col)
                    max_length = 0
                    for cell in worksheet[column_letter]:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            logging.info(f"엑셀 파일로 저장 완료: {file_path}")

        except Exception as e:
            logging.error(f"엑셀 파일 저장 실패: {e}")

if __name__ == "__main__":
    asyncio.run(crawl_kfb_loan_rates())