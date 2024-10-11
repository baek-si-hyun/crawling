import asyncio
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import pandas as pd
import os
import urllib.parse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook
import html  # HTML 엔티티 디코딩을 위해 추가

async def main():
    data_list = []

    # 검색어 입력 받기
    search_keyword = input("검색어를 입력하세요 (없으면 엔터): ").strip()

    # 페이지 수 설정 (한 페이지당 5개씩)
    select_page_num = 5  # 페이지 번호는 1부터 시작하므로 1~5페이지를 의미합니다.

    # 검색어에 따라 Excel 파일 이름 설정
    if search_keyword:
        encoded_keyword = urllib.parse.quote(search_keyword)
        excel_file = f'기획재정부_{search_keyword}.xlsx'
    else:
        excel_file = '기획재정부.xlsx'

    # 기존 데이터 로드
    if os.path.exists(excel_file):
        existing_df = pd.read_excel(excel_file)
    else:
        existing_df = pd.DataFrame(columns=['title', 'depart', 'date', 'download_link', 'preview_link'])

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)  # GUI를 표시하지 않음
        page = await browser.new_page()

        # 1단계: 목록 페이지에서 데이터 수집
        for page_num in range(1, select_page_num + 1):
            if search_keyword:
                if page_num == 1:
                    # 첫 번째 페이지 URL (검색어 포함, pageIndex 없음)
                    url = f"https://www.moef.go.kr/nw/nes/nesdta.do?searchBbsId1=MOSFBBS_000000000028&menuNo=4010100&searchKeyword3={encoded_keyword}&searchCondition3=0&searchSilDeptId1=&kwd1="
                else:
                    # 두 번째 페이지부터 URL (검색어 및 pageIndex 포함)
                    url = f"https://www.moef.go.kr/nw/nes/nesdta.do?searchBbsId1=MOSFBBS_000000000028&menuNo=4010100&pageIndex={page_num}&searchKeyword3={encoded_keyword}&searchCondition3=0&searchSilDeptId1=&kwd1="
            else:
                # 검색어가 없을 때의 URL
                url = f"https://www.moef.go.kr/nw/nes/nesdta.do?searchBbsId1=MOSFBBS_000000000028&menuNo=4010100&pageIndex={page_num}"

            print(f"페이지 {page_num} 크롤링 중: {url}")
            try:
                await page.goto(url)
                await page.wait_for_timeout(1000)  # 1초 지연

                # 페이지의 HTML 가져오기
                content = await page.content()
                soup = BeautifulSoup(content, 'html.parser')

                # 리스트 아이템 선택
                list_items = soup.find_all('li')

                # 페이지 내 데이터 수 카운트 초기화
                page_data_count = 0

                for item in list_items:
                    # 제목과 href를 포함한 a 태그 찾기
                    h3_tag = item.find('h3')
                    if h3_tag:
                        a_tag = h3_tag.find('a')
                        if a_tag:
                            title = a_tag.get_text(strip=True)

                            # href 추출 및 절대 URL로 변환 (공백 제거, HTML 엔티티 디코딩)
                            href = a_tag.get('href', '').strip()
                            href = html.unescape(href)  # HTML 엔티티 디코딩
                            if href.startswith('/'):
                                href = 'https://www.moef.go.kr' + href
                            elif href.startswith('dtl.jsp'):
                                href = 'https://www.moef.go.kr/' + href
                            else:
                                href = 'https://www.moef.go.kr' + href

                            # 날짜 추출
                            date_span = item.find('span', class_='date')
                            if date_span:
                                date = date_span.get_text(strip=True)
                            else:
                                date = None

                            # 과 추출
                            depart_span = item.find('span', class_='depart')
                            if depart_span:
                                depart = depart_span.get_text(strip=True)
                            else:
                                depart = None

                            # 첨부파일 다운로드 링크 추출
                            download_link_tag = item.find('a', class_='icoFile fileDown')
                            if download_link_tag:
                                download_link = download_link_tag.get('href', '').strip()
                                download_link = html.unescape(download_link)  # HTML 엔티티 디코딩
                                if download_link.startswith('/'):
                                    download_link = 'https://www.moef.go.kr' + download_link
                                elif download_link.startswith('dtl.jsp'):
                                    download_link = 'https://www.moef.go.kr/' + download_link
                                else:
                                    download_link = 'https://www.moef.go.kr' + download_link
                            else:
                                download_link = None

                            # 첨부파일 미리보기 링크 추출
                            preview_link_tag = item.find('a', class_='icoFile fileView')
                            if preview_link_tag:
                                preview_link = preview_link_tag.get('href', '').strip()
                                preview_link = html.unescape(preview_link)  # HTML 엔티티 디코딩
                                if preview_link.startswith('/'):
                                    preview_link = 'https://www.moef.go.kr' + preview_link
                                elif preview_link.startswith('dtl.jsp'):
                                    preview_link = 'https://www.moef.go.kr/' + preview_link
                                else:
                                    preview_link = 'https://www.moef.go.kr' + preview_link
                            else:
                                preview_link = None

                            data = {
                                'title': title,
                                'depart': depart,
                                'date': date,
                                'download_link': download_link,
                                'preview_link': preview_link
                            }

                            data_list.append(data)
                            page_data_count += 1

                # 데이터가 하나도 수집되지 않은 페이지가 나오면 반복문 종료
                if page_data_count == 0:
                    print(f"페이지 {page_num}에 데이터가 없습니다. 더 이상의 페이지가 없다고 판단하여 크롤링을 종료합니다.")
                    break

            except Exception as e:
                print(f"페이지 {page_num} 크롤링 중 오류 발생: {e}")
                continue

        await browser.close()

        # 데이터가 수집되지 않은 경우 Excel 저장 단계 건너뛰기
        if not data_list:
            print("\n크롤링한 데이터가 없습니다. Excel 파일을 생성하지 않습니다.")
            return

        # 2단계: 데이터 저장 (Excel 파일로) 및 열 너비 조정
        try:
            # 새로운 데이터프레임 생성
            new_df = pd.DataFrame(data_list)

            # 기존 데이터와 새로운 데이터 병합
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)

            # 중복 제거 ('title'과 'date'가 모두 동일한 경우 중복으로 간주)
            combined_df.drop_duplicates(subset=['title', 'date'], keep='first', inplace=True)

            # Excel 파일로 저장 (열 너비 자동 조정)
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # 각 열의 최대 길이를 계산하여 열 너비를 설정
                for i, column in enumerate(combined_df.columns, 1):
                    # 열의 최대 길이 계산
                    max_length = combined_df[column].astype(str).map(len).max()
                    # 열 제목의 길이와 비교하여 더 큰 값 선택
                    max_length = max(max_length, len(column)) + 2  # 여유 공간 추가
                    # 열 번호를 열 문자로 변환
                    column_letter = get_column_letter(i)
                    # 열 너비 설정
                    worksheet.column_dimensions[column_letter].width = max_length

            print(f"\n새로운 데이터 {len(new_df)}개를 수집하였습니다.")
            print(f"중복을 제거한 후 총 {len(combined_df)}개의 데이터가 '{excel_file}'에 저장되었습니다.")

            # 하이퍼링크 설정을 위해 워크북 다시 로드
            wb = load_workbook(excel_file)
            ws = wb['Sheet1']

            # 'download_link'와 'preview_link' 컬럼 인덱스 찾기
            try:
                download_col = combined_df.columns.get_loc('download_link') + 1  # 1부터 시작
                preview_col = combined_df.columns.get_loc('preview_link') + 1
            except ValueError as ve:
                print(f"컬럼을 찾을 수 없습니다: {ve}")
                return

            # 하이퍼링크 설정
            for row in range(2, len(combined_df) + 2):  # 헤더 제외
                # 다운로드 링크 하이퍼링크 설정
                download_cell = ws.cell(row=row, column=download_col)
                if download_cell.value and isinstance(download_cell.value, str) and download_cell.value.startswith('http'):
                    download_cell.hyperlink = download_cell.value
                    download_cell.font = Font(color='0000FF', underline='single')

                # 미리보기 링크 하이퍼링크 설정
                preview_cell = ws.cell(row=row, column=preview_col)
                if preview_cell.value and isinstance(preview_cell.value, str) and preview_cell.value.startswith('http'):
                    preview_cell.hyperlink = preview_cell.value
                    preview_cell.font = Font(color='0000FF', underline='single')

            # 워크북 저장
            wb.save(excel_file)

            print(f"\n하이퍼링크가 추가된 Excel 파일 '{excel_file}'을 생성하였습니다.")

        except Exception as e:
            print(f"엑셀 파일 저장 중 오류 발생: {e}")

if __name__ == "__main__":
    asyncio.run(main())