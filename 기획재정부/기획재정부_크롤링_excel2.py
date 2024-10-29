import asyncio
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import pandas as pd
import os
import urllib.parse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook
import html  # HTML 엔티티 디코딩을 위해 추가
import datetime  # 날짜 및 시간 처리를 위해 추가
import logging
import re  # 정규표현식을 위해 추가

# 로깅 설정
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')


def format_date(input_date):
    """
    사용자가 입력한 날짜를 'yyyy.mm.dd.' 형식으로 변환합니다.
    입력 형식이 'yyyymmdd'인 경우 'yyyy.mm.dd.'로 변환합니다.
    이미 'yyyy.mm.dd' 형식인 경우 마지막에 '.'을 추가합니다.
    유효하지 않은 형식일 경우 빈 문자열을 반환합니다.
    """
    if not input_date:
        return ""

    # 'yyyy.mm.dd' 형식 확인
    if re.match(r'^\d{4}\.\d{2}\.\d{2}$', input_date):
        return input_date + '.'  # 마지막에 '.' 추가

    # 'yyyymmdd' 형식 확인
    elif re.match(r'^\d{8}$', input_date):
        return f"{input_date[:4]}.{input_date[4:6]}.{input_date[6:8]}."

    else:
        # 유효하지 않은 형식일 경우 빈 문자열 반환
        logging.warning(f"날짜 형식이 올바르지 않습니다: {input_date}")
        return ""


async def main():
    data_list = []

    # 검색어 및 날짜 입력 받기
    search_keyword = input("검색어를 입력하세요 (없으면 엔터): ").strip()
    search_start_date_input = input("시작일을 입력하세요 (형식: yyyy.mm.dd 또는 yyyymmdd, 없으면 엔터): ").strip()
    search_end_date_input = input("마감일을 입력하세요 (형식: yyyy.mm.dd 또는 yyyymmdd, 없으면 엔터): ").strip()

    # 날짜 형식 변환
    search_start_date = format_date(search_start_date_input)
    search_end_date = format_date(search_end_date_input)

    if search_start_date_input and not search_start_date:
        logging.warning(f"시작일 형식이 올바르지 않습니다: {search_start_date_input}")
    if search_end_date_input and not search_end_date:
        logging.warning(f"마감일 형식이 올바르지 않습니다: {search_end_date_input}")

    # 현재 날짜와 시간을 포함한 Excel 파일 이름 설정
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    if search_keyword:
        # 검색어에 특수문자가 있을 경우 파일명에서 문제가 발생할 수 있으므로, 이를 제거하거나 대체할 수 있습니다.
        safe_keyword = re.sub(r'[\\/*?:"<>|]', "_", search_keyword)
        excel_file = f'기획재정부_{safe_keyword}_{timestamp}.xlsx'
    else:
        excel_file = f'기획재정부_{timestamp}.xlsx'

    # 기존 데이터 로드
    if os.path.exists(excel_file):
        existing_df = pd.read_excel(excel_file)
        logging.info(f"기존 엑셀 파일 '{excel_file}'을 로드하였습니다.")
    else:
        existing_df = pd.DataFrame(columns=['title', 'depart', 'date', 'preview_link', 'direct_link'])
        logging.info(f"새로운 데이터프레임을 생성하였습니다.")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)  # 디버깅을 위해 headless=False로 설정
        page = await browser.new_page()

        # 대상 웹사이트 기본 URL
        base_url = "https://www.moef.go.kr/nw/nes/nesdta.do?searchBbsId1=MOSFBBS_000000000028&menuNo=4010100"

        logging.info(f"페이지 접속: {base_url}")
        try:
            await page.goto(base_url, timeout=60000)  # 60초 타임아웃
        except Exception as e:
            logging.error(f"페이지 접속 중 오류 발생: {e}")
            await browser.close()
            return

        # 페이지가 완전히 로드될 때까지 대기
        await page.wait_for_load_state('networkidle')

        # 검색어 및 날짜가 있는 경우, 입력 필드에 값 채우기
        try:
            if search_keyword:
                # 검색어 입력
                search_input_selector = 'input#searchKeyword1'
                await page.fill(search_input_selector, search_keyword)
                logging.info(f"검색어 '{search_keyword}'를 입력하였습니다.")

            if search_start_date:
                # 시작일 입력
                search_start_selector = 'input#searchStartDt'
                await page.fill(search_start_selector, search_start_date)
                logging.info(f"시작일 '{search_start_date}'를 입력하였습니다.")

            if search_end_date:
                # 마감일 입력
                search_end_selector = 'input#searchEndDt'
                await page.fill(search_end_selector, search_end_date)
                logging.info(f"마감일 '{search_end_date}'를 입력하였습니다.")

            if search_keyword or search_start_date or search_end_date:
                # 검색 버튼 클릭
                search_button_selector = 'input[type="submit"][title="검색하기"]'
                await page.click(search_button_selector)
                logging.info("검색 버튼을 클릭하였습니다.")

                # 검색 결과가 로드될 때까지 대기
                try:
                    await page.wait_for_selector('ul.boardType3.explnList', timeout=60000)  # 60초 타임아웃
                    logging.info("검색 결과가 로드되었습니다.")
                except Exception as e:
                    logging.error(f"검색 결과 로드 중 오류 발생: {e}")
                    # 검색 결과가 없음을 나타내는 요소가 있는지 확인 (실제 셀렉터로 수정 필요)
                    no_result_selector = 'div.noResult'  # 실제 셀렉터로 수정 필요
                    if await page.query_selector(no_result_selector):
                        logging.info("검색 결과가 없습니다.")
                    else:
                        logging.error("알 수 없는 오류가 발생했습니다.")
                    # 오류 발생 시 스크린샷 저장
                    await page.screenshot(path="error_search_result.png")
                    logging.info("오류 발생 시 스크린샷을 저장하였습니다: error_search_result.png")
                    await browser.close()
                    return
        except Exception as e:
            logging.error(f"검색어 또는 날짜 입력, 검색 버튼 클릭 중 오류 발생: {e}")
            # 오류 발생 시 스크린샷 저장
            await page.screenshot(path="error_search.png")
            logging.info("오류 발생 시 스크린샷을 저장하였습니다: error_search.png")
            await browser.close()
            return

        # 1단계: 페이지네이션을 통해 모든 데이터를 수집
        visited_pages = set()
        while True:
            # 현재 페이지의 HTML 가져오기
            content = await page.content()
            soup = BeautifulSoup(content, 'html.parser')

            # 현재 페이지 번호 추출
            current_page_element = await page.query_selector('div.boardPage span.num strong')
            if current_page_element:
                current_page_text = await current_page_element.inner_text()
                current_page = int(current_page_text.strip())
            else:
                logging.warning("현재 페이지 번호를 찾을 수 없습니다.")
                current_page = None

            # 현재 페이지가 이미 방문한 페이지인지 확인
            if current_page is not None and current_page in visited_pages:
                logging.info(f"페이지 {current_page}은 이미 방문하였습니다. 크롤링을 종료합니다.")
                break
            elif current_page is not None:
                visited_pages.add(current_page)

            # 리스트 아이템 선택 및 데이터 수집
            list_items = soup.select('ul.boardType3.explnList > li')
            if not list_items:
                logging.info("리스트 아이템이 없습니다. 크롤링을 종료합니다.")
                break

            page_data_count = 0

            for item in list_items:
                # 제목과 href를 포함한 a 태그 찾기
                h3_tag = item.find('h3')
                if h3_tag:
                    a_tag = h3_tag.find('a')
                    if a_tag and 'href' in a_tag.attrs:
                        title = a_tag.get_text(strip=True)

                        # href에서 고유 ID 추출
                        href = a_tag['href']
                        unique_id_match = re.search(r"fn_egov_select\('(.+)'\)", href)
                        if unique_id_match:
                            unique_id = unique_id_match.group(1)
                            # 상세보기 링크 생성
                            detail_link = f"https://www.moef.go.kr/nw/nes/detailNesDtaView.do?searchBbsId1=MOSFBBS_000000000028&searchNttId1={unique_id}&menuNo=4010100"
                        else:
                            logging.warning(f"고유 ID를 추출할 수 없습니다: {href}")
                            detail_link = ""

                        # 날짜 추출
                        date_span = item.find('span', class_='date')
                        if date_span:
                            date = date_span.get_text(strip=True)
                        else:
                            date = None

                        # 부서 추출
                        depart_span = item.find('span', class_='depart')
                        if depart_span:
                            depart = depart_span.get_text(strip=True)
                        else:
                            depart = None

                        # 첨부파일 미리보기 링크 추출
                        preview_link_tag = item.find('a', class_='icoFile fileView')
                        if preview_link_tag and 'href' in preview_link_tag.attrs:
                            preview_link = preview_link_tag['href']
                            preview_link = urllib.parse.urljoin("https://www.moef.go.kr", preview_link)
                        else:
                            preview_link = None

                        data = {
                            'title': title,
                            'depart': depart,
                            'date': date,
                            'preview_link': preview_link,
                            'direct_link': detail_link  # 상세보기 링크를 direct_link로 사용
                        }

                        data_list.append(data)
                        page_data_count += 1

            logging.info(f"페이지 {current_page}에서 {page_data_count}개의 데이터를 수집하였습니다.")

            # 페이지 번호 링크들을 수집 (현재 페이지 포함)
            pagination_elements = await page.query_selector_all('div.boardPage span.num')
            page_indices = []

            for elem in pagination_elements:
                # 현재 페이지
                strong_elem = await elem.query_selector('strong')
                if strong_elem:
                    page_num_text = await strong_elem.inner_text()
                    page_num = int(page_num_text.strip())
                    page_indices.append(page_num)
                else:
                    # 다른 페이지 링크
                    a_elem = await elem.query_selector('a')
                    if a_elem:
                        page_num_text = await a_elem.inner_text()
                        page_num = int(page_num_text.strip())
                        page_indices.append(page_num)

            # 현재 페이지 세트에서 방문하지 않은 페이지 번호들을 찾음
            pages_to_visit = [p for p in page_indices if p not in visited_pages]

            # 방문하지 않은 페이지로 이동
            if pages_to_visit:
                next_page_num = pages_to_visit[0]
                logging.info(f"페이지 {next_page_num}로 이동합니다.")

                # 페이지 번호 클릭
                await page.evaluate(f"fn_egov_link_page({next_page_num});")

                # 페이지 로드 대기
                await page.wait_for_load_state('networkidle')
                try:
                    await page.wait_for_selector('ul.boardType3.explnList', timeout=60000)
                except Exception as e:
                    logging.error(f"페이지 {next_page_num} 로드 중 오류 발생: {e}")
                    await page.screenshot(path=f"error_page_load_{next_page_num}.png")
                    logging.info(f"페이지 {next_page_num} 로드 오류 시 스크린샷을 저장하였습니다: error_page_load_{next_page_num}.png")
                    break  # 또는 continue로 다음 페이지 시도
            else:
                # 현재 페이지 세트의 모든 페이지를 방문했으므로 '다음' 버튼 확인
                next_button = await page.query_selector('div.boardPage span.next a')
                if next_button:
                    # 'Next' 버튼의 href에서 이동할 페이지 번호를 추출하여 이미 방문한 페이지인지 확인
                    onclick_attr = await next_button.get_attribute('onclick')
                    match = re.search(r"fn_egov_link_page\((\d+)\)", onclick_attr)
                    if match:
                        next_page_num = int(match.group(1))
                        if next_page_num in visited_pages:
                            logging.info(f"페이지 {next_page_num}은 이미 방문하였습니다. 크롤링을 종료합니다.")
                            break
                        else:
                            await next_button.click()
                            await page.wait_for_load_state('networkidle')
                            logging.info(f"'다음' 버튼 클릭하여 페이지 {next_page_num}로 이동하였습니다.")
                            # 페이지 로드 대기
                            try:
                                await page.wait_for_selector('ul.boardType3.explnList', timeout=60000)
                            except Exception as e:
                                logging.error(f"'다음' 버튼 클릭 후 페이지 로드 중 오류 발생: {e}")
                                await page.screenshot(path="error_next_page.png")
                                logging.info("오류 발생 시 스크린샷을 저장하였습니다: error_next_page.png")
                                break
                    else:
                        logging.warning("다음 페이지 번호를 추출할 수 없습니다. 크롤링을 종료합니다.")
                        break
                else:
                    logging.info("더 이상의 '다음' 버튼이 없습니다. 크롤링을 종료합니다.")
                    break

        await browser.close()

        # 데이터가 수집되지 않은 경우 Excel 저장 단계 건너뛰기
        if not data_list:
            logging.warning("크롤링한 데이터가 없습니다. Excel 파일을 생성하지 않습니다.")
            return

        # 2단계: 데이터 저장 (Excel 파일로) 및 열 너비 조정
        try:
            # 새로운 데이터프레임 생성
            new_df = pd.DataFrame(data_list)

            # 기존 데이터와 새로운 데이터 병합
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)

            # 중복 제거 ('title'과 'date'가 모두 동일한 경우 중복으로 간주)
            combined_df.drop_duplicates(subset=['title', 'date'], keep='first', inplace=True)

            # Excel 파일로 저장 (열 너비 자동 조정)
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # 각 열의 최대 길이를 계산하여 열 너비를 설정
                for i, column in enumerate(combined_df.columns, 1):
                    # 열의 최대 길이 계산
                    max_length = combined_df[column].astype(str).map(len).max()
                    # 열 제목의 길이와 비교하여 더 큰 값 선택
                    max_length = max(max_length, len(column)) + 2  # 여유 공간 추가
                    # 열 번호를 열 문자로 변환
                    column_letter = get_column_letter(i)
                    # 열 너비 설정
                    worksheet.column_dimensions[column_letter].width = max_length

            logging.info(f"새로운 데이터 {len(new_df)}개를 수집하였습니다.")
            logging.info(f"중복을 제거한 후 총 {len(combined_df)}개의 데이터가 '{excel_file}'에 저장되었습니다.")

            # 하이퍼링크 설정을 위해 워크북 다시 로드
            wb = load_workbook(excel_file)
            ws = wb['Sheet1']

            # 'direct_link' 컬럼 인덱스 찾기
            try:
                direct_link_col = combined_df.columns.get_loc('direct_link') + 1  # 1부터 시작
            except ValueError as ve:
                logging.error(f"'direct_link' 컬럼을 찾을 수 없습니다: {ve}")
                return

            # 하이퍼링크 설정
            for row in range(2, len(combined_df) + 2):  # 헤더 제외
                # 직접 링크 하이퍼링크 설정
                direct_link_cell = ws.cell(row=row, column=direct_link_col)
                if direct_link_cell.value and isinstance(direct_link_cell.value, str) and direct_link_cell.value.startswith('http'):
                    direct_link_cell.hyperlink = direct_link_cell.value
                    direct_link_cell.font = Font(color='0000FF', underline='single')

            # 워크북 저장
            wb.save(excel_file)

            logging.info(f"하이퍼링크가 추가된 Excel 파일 '{excel_file}'을 생성하였습니다.")

        except Exception as e:
            logging.error(f"엑셀 파일 저장 중 오류 발생: {e}")


if __name__ == "__main__":
    asyncio.run(main())