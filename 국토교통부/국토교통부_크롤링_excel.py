import asyncio
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import pandas as pd
import os
import urllib.parse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook
import html  # HTML 엔티티 디코딩을 위해 추가
from datetime import datetime, timedelta

async def main():
    data_list = []

    # 검색어 입력 받기
    search_keyword = input("검색어를 입력하세요 (없으면 엔터): ").strip()

    # 페이지 수 설정
    max_page_num = 5  # 1부터 5페이지까지

    # 오늘 날짜와 1년 전 날짜 계산
    today = datetime.today()
    one_year_ago = today - timedelta(days=365)

    # 날짜를 YYYY-MM-DD 형식으로 포맷팅
    search_regdate_s = one_year_ago.strftime('%Y-%m-%d')
    search_regdate_e = today.strftime('%Y-%m-%d')

    # 검색어에 따라 Excel 파일 이름 설정
    if search_keyword:
        encoded_keyword = urllib.parse.quote(search_keyword)
        excel_file = f'국토교통부_{search_keyword}.xlsx'
    else:
        excel_file = '국토교통부.xlsx'

    # 기존 데이터 로드
    if os.path.exists(excel_file):
        existing_df = pd.read_excel(excel_file)
    else:
        existing_df = pd.DataFrame(columns=['title', 'category', 'date', 'download_link', 'preview_link'])

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)  # GUI를 표시하지 않음
        page = await browser.new_page()

        # 1단계: 목록 페이지에서 데이터 수집
        for page_num in range(1, max_page_num + 1):
            if search_keyword:
                # 검색어가 있을 경우 URL 구성
                url = (
                    f"https://www.molit.go.kr/USR/NEWS/m_71/lst.jsp?"
                    f"search_section=&search={encoded_keyword}&search_gubun=1&search_gubun1=all&"
                    f"srch_usr_titl=Y&srch_usr_ctnt=&psize=10&search_regdate_s={search_regdate_s}&"
                    f"search_regdate_e={search_regdate_e}&srch_cate=&srch_dept_nm=&search_kind=&"
                    f"search_gbn=&lst_gbn=T&lcmspage={page_num}"
                )
            else:
                # 검색어가 없을 경우 URL 구성
                url = (
                    f"https://www.molit.go.kr/USR/NEWS/m_71/lst.jsp?"
                    f"search_section=&search=&search_gubun=1&search_gubun1=all&"
                    f"srch_usr_titl=Y&srch_usr_ctnt=&psize=10&search_regdate_s={search_regdate_s}&"
                    f"search_regdate_e={search_regdate_e}&srch_cate=&srch_dept_nm=&search_kind=&"
                    f"search_gbn=&lst_gbn=T&lcmspage={page_num}"
                )

            print(f"페이지 {page_num} 크롤링 중: {url}")
            try:
                await page.goto(url)
                # await page.wait_for_timeout(1000)  # 1초 지연

                # 페이지의 HTML 가져오기
                content = await page.content()
                soup = BeautifulSoup(content, 'html.parser')

                # tbody 내의 tr 태그 선택
                tbody = soup.find('tbody')
                if not tbody:
                    print(f"페이지 {page_num}에서 tbody를 찾을 수 없습니다.")
                    continue

                rows = tbody.find_all('tr')

                # 페이지 내 데이터 수 카운트 초기화
                page_data_count = 0

                for row in rows:
                    # 제목, 날짜, href 추출
                    title_td = row.find('td', class_='bd_title')
                    date_td = row.find('td', class_='bd_date')
                    field_td = row.find('td', class_='bd_field')

                    if title_td and date_td:
                        a_tag = title_td.find('a')
                        if a_tag:
                            # 제목 추출 (a 태그 내의 텍스트)
                            title = a_tag.get_text(strip=True)

                            # href 추출 및 절대 URL로 변환 (공백 제거, HTML 엔티티 디코딩)
                            href = a_tag.get('href', '').strip()
                            href = html.unescape(href)  # HTML 엔티티 디코딩
                            if href.startswith('/'):
                                href = 'https://www.molit.go.kr' + href
                            elif href.startswith('dtl.jsp'):
                                href = 'https://www.molit.go.kr/USR/NEWS/m_71/' + href
                            else:
                                href = 'https://www.molit.go.kr/USR/NEWS/m_71/' + href  # 기본적으로 상대경로 처리

                            # 날짜 추출
                            date = date_td.get_text(strip=True)

                            # 카테고리 추출
                            category = field_td.get_text(strip=True) if field_td else None

                            # 다운로드 링크와 미리보기 링크는 추후 상세 페이지에서 추출
                            data = {
                                'title': title,
                                'category': category,
                                'date': date,
                                'href': href,  # 상세 페이지 URL
                                'download_link': None,
                                'preview_link': None
                            }

                            data_list.append(data)
                            page_data_count += 1

                # 데이터가 하나도 수집되지 않은 페이지가 나오면 반복문 종료
                if page_data_count == 0:
                    print(f"페이지 {page_num}에 데이터가 없습니다. 더 이상의 페이지가 없다고 판단하여 크롤링을 종료합니다.")
                    break

            except Exception as e:
                print(f"페이지 {page_num} 크롤링 중 오류 발생: {e}")
                continue

        # 목록 페이지 크롤링이 끝났으므로 브라우저 닫기
        await browser.close()

        # 데이터가 수집되지 않은 경우 Excel 저장 단계 건너뛰기
        if not data_list:
            print("\n크롤링한 데이터가 없습니다. Excel 파일을 생성하지 않습니다.")
            return

        # 2단계: 상세 페이지에서 다운로드 링크와 미리보기 링크 수집
        print("\n상세 페이지에서 다운로드 링크와 미리보기 링크를 수집 중입니다...\n")

        # 상세 페이지 수집을 위해 브라우저 재실행
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()

        for idx, entry in enumerate(data_list, start=1):
            href = entry.get('href')  # 'href' 키 안전하게 접근
            if not href:
                print(f"상세 페이지 {idx}에 href가 없습니다. 건너뜁니다.")
                continue

            print(f"상세 페이지 {idx}/{len(data_list)}: {href}")
            try:
                await page.goto(href)
                await page.wait_for_timeout(1000)  # 1초 지연

                # 상세 페이지의 HTML 가져오기
                detail_content = await page.content()
                detail_soup = BeautifulSoup(detail_content, 'html.parser')

                # <li class="file"> 찾기
                li_file = detail_soup.find('li', class_='file')
                if li_file:
                    a_tags = li_file.find_all('a')
                    download_link = None
                    preview_link = None

                    for a in a_tags:
                        href_attr = a.get('href', '').strip()
                        href_attr = html.unescape(href_attr)  # HTML 엔티티 디코딩
                        title_attr = a.get('title', '').strip()

                        # 다운로드 링크 찾기
                        if 'download' in href_attr or href_attr.endswith(('.pdf', '.hwpx', '.xlsx', '.docx')):
                            if href_attr.startswith('/'):
                                download_link = 'https://www.molit.go.kr' + href_attr
                            elif href_attr.startswith('dtl.jsp'):
                                download_link = 'https://www.molit.go.kr/USR/NEWS/m_71/' + href_attr
                            else:
                                download_link = 'https://www.molit.go.kr' + href_attr
                        # 미리보기 링크 찾기
                        elif title_attr == '첨부파일 미리보기':
                            if href_attr.startswith('/'):
                                preview_link = 'https://www.molit.go.kr' + href_attr
                            else:
                                preview_link = href_attr  # 절대경로가 아닐 경우 그대로 저장

                    # 추출된 링크를 데이터에 추가
                    data_list[idx - 1]['download_link'] = download_link
                    data_list[idx - 1]['preview_link'] = preview_link
                else:
                    # <li class="file">이 없으면 링크 없음
                    data_list[idx - 1]['download_link'] = None
                    data_list[idx - 1]['preview_link'] = None

            except Exception as e:
                print(f"상세 페이지 {href} 크롤링 중 오류 발생: {e}")
                data_list[idx - 1]['download_link'] = None
                data_list[idx - 1]['preview_link'] = None
                continue

        # 상세 페이지 크롤링이 끝났으므로 브라우저 닫기
        await browser.close()

    # 3단계: 데이터 저장 (Excel 파일로) 및 열 너비 조정 및 하이퍼링크 설정
        try:
            # 'href'를 제외한 데이터만 추출
            data_without_href = [
                {
                    'title': d['title'],
                    'category': d['category'],
                    'date': d['date'],
                    'download_link': d['download_link'],
                    'preview_link': d['preview_link']
                }
                for d in data_list
            ]

            # 새로운 데이터프레임 생성
            new_df = pd.DataFrame(data_without_href)

            # 기존 데이터와 새로운 데이터 병합
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)

            # 중복 제거 ('title'과 'date'가 모두 동일한 경우 중복으로 간주)
            combined_df.drop_duplicates(subset=['title', 'date'], keep='first', inplace=True)

            # Excel 파일로 저장 (열 너비 자동 조정)
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # 각 열의 최대 길이를 계산하여 열 너비를 설정
                for i, column in enumerate(combined_df.columns, 1):
                    # 열의 최대 길이 계산
                    max_length = combined_df[column].astype(str).map(len).max()
                    # 열 제목의 길이와 비교하여 더 큰 값 선택
                    max_length = max(max_length, len(column)) + 2  # 여유 공간 추가
                    # 열 번호를 열 문자로 변환
                    column_letter = get_column_letter(i)
                    # 열 너비 설정
                    worksheet.column_dimensions[column_letter].width = max_length

            print(f"\n새로운 데이터 {len(new_df)}개를 수집하였습니다.")
            print(f"중복을 제거한 후 총 {len(combined_df)}개의 데이터가 '{excel_file}'에 저장되었습니다.")

            # 하이퍼링크 설정을 위해 워크북 다시 로드
            wb = load_workbook(excel_file)
            ws = wb['Sheet1']

            # 'download_link'와 'preview_link' 컬럼 인덱스 찾기
            try:
                download_col = combined_df.columns.get_loc('download_link') + 1  # 1부터 시작
                preview_col = combined_df.columns.get_loc('preview_link') + 1
            except ValueError as ve:
                print(f"컬럼을 찾을 수 없습니다: {ve}")
                return

            # 하이퍼링크 설정
            for row in range(2, len(combined_df) + 2):  # 헤더 제외
                # 다운로드 링크 하이퍼링크 설정
                download_cell = ws.cell(row=row, column=download_col)
                if download_cell.value and isinstance(download_cell.value, str) and download_cell.value.startswith('http'):
                    download_cell.hyperlink = download_cell.value
                    download_cell.font = Font(color='0000FF', underline='single')

                # 미리보기 링크 하이퍼링크 설정
                preview_cell = ws.cell(row=row, column=preview_col)
                if preview_cell.value and isinstance(preview_cell.value, str) and preview_cell.value.startswith('http'):
                    preview_cell.hyperlink = preview_cell.value
                    preview_cell.font = Font(color='0000FF', underline='single')

            # 워크북 저장
            wb.save(excel_file)

            print(f"\n하이퍼링크가 추가된 Excel 파일 '{excel_file}'을 생성하였습니다.")

        except Exception as e:
            print(f"엑셀 파일 저장 중 오류 발생: {e}")

if __name__ == "__main__":
    asyncio.run(main())