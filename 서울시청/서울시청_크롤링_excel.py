import asyncio
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import pandas as pd
import os
import urllib.parse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook
from datetime import datetime, timedelta


async def main():
    # 데이터 저장을 위한 리스트 초기화
    data_list = []

    # 검색어 입력 받기
    search_keyword = input("검색어를 입력하세요 (없으면 엔터): ").strip()

    # 페이지 수 설정
    max_page_num = 5  # 1부터 5페이지까지

    # 오늘 날짜와 1년 전 날짜 계산 (YYYYMMDD 형식)
    today = datetime.today().strftime('%Y%m%d')
    one_year_ago = (datetime.today() - timedelta(days=365)).strftime('%Y%m%d')

    # 검색어에 따라 Excel 파일 이름 설정
    if search_keyword:
        excel_file = f'서울시청_{search_keyword}.xlsx'
    else:
        excel_file = '서울시청.xlsx'

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()

        # 1단계: 목록 페이지에서 데이터 수집
        for page_num in range(1, max_page_num + 1):
            if search_keyword:
                # 검색어가 있을 경우 URL 구성
                url = f"https://www.seoul.go.kr/news/news_report.do#list/{page_num}/srchBeginDt={one_year_ago}&srchEndDt={today}&cntPerPage=10&srchKey=sj&srchText={urllib.parse.quote(search_keyword)}"
            else:
                # 검색어가 없을 경우 URL 구성
                url = f"https://www.seoul.go.kr/news/news_report.do#list/{page_num}/srchBeginDt={one_year_ago}&srchEndDt={today}&cntPerPage=10"

            print(f"페이지 {page_num} 크롤링 중: {url}")
            try:
                # 페이지 이동
                await page.goto(url, wait_until='networkidle')
                # 필요한 요소가 로드될 때까지 대기
                await page.wait_for_selector('tbody')

                # 페이지의 HTML 가져오기
                content = await page.content()
                soup = BeautifulSoup(content, 'html.parser')

                # tbody 내의 tr 태그 선택
                tbody = soup.find('tbody')
                if not tbody:
                    print(f"페이지 {page_num}에서 tbody를 찾을 수 없습니다.")
                    continue

                rows = tbody.find_all('tr')

                # 페이지 내 데이터 수 카운트 초기화
                page_data_count = 0

                for row in rows:
                    # 제목, 담당부서, 날짜, data-code 추출
                    a_tag = row.find('a', attrs={'data-code': True})
                    if a_tag:
                        title = a_tag.get_text(strip=True)
                        data_code = a_tag.get('data-code', '').strip()
                        if not data_code:
                            continue

                        # 담당부서 추출
                        tds = row.find_all('td')
                        department_td = tds[2] if len(tds) > 2 else None
                        department = department_td.get_text(
                            strip=True) if department_td else ''

                        # 날짜 추출
                        date_td = tds[3] if len(tds) > 3 else None
                        date = date_td.get_text(strip=True) if date_td else ''

                        # 상세 페이지 URL 구성
                        href = f"https://www.seoul.go.kr/news/news_report.do#view/{data_code}"

                        # 데이터 저장
                        data = {
                            'title': title,
                            'department': department,
                            'date': date,
                            'data_code': data_code,
                            'href': href,
                            'download_link': None,
                            'preview_link': None
                        }

                        data_list.append(data)
                        page_data_count += 1

                # 데이터가 하나도 수집되지 않은 페이지가 나오면 반복문 종료
                if page_data_count == 0:
                    print(
                        f"페이지 {page_num}에 데이터가 없습니다. 더 이상의 페이지가 없다고 판단하여 크롤링을 종료합니다.")
                    break

            except Exception as e:
                print(f"페이지 {page_num} 크롤링 중 오류 발생: {e}")
                continue

        # 목록 페이지 크롤링이 끝났으므로 브라우저 닫기
        await browser.close()

        # 데이터가 수집되지 않은 경우 Excel 저장 단계 건너뛰기
        if not data_list:
            print("\n크롤링한 데이터가 없습니다. Excel 파일을 생성하지 않습니다.")
            return

        # 2단계: 상세 페이지에서 다운로드 링크와 미리보기 링크 수집
        print("\n상세 페이지에서 다운로드 링크와 미리보기 링크를 수집 중입니다...\n")

        # 상세 페이지 수집을 위해 브라우저 재실행
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()

        for idx, entry in enumerate(data_list, start=1):
            href = entry['href']
            data_code = entry['data_code']
            print(f"상세 페이지 {idx}/{len(data_list)}: {href}")
            try:
                # 페이지 이동
                await page.goto(href, wait_until='networkidle')
                # 필요한 요소가 로드될 때까지 대기
                await page.wait_for_selector('thead')

                # 상세 페이지의 HTML 가져오기
                detail_content = await page.content()
                detail_soup = BeautifulSoup(detail_content, 'html.parser')

                # 첨부파일 정보 추출
                download_links = []
                preview_links = []

                # thead 내의 첨부파일 관련 tr 태그 찾기
                tr_tags = detail_soup.select('thead tr')

                for tr in tr_tags:
                    # 다운로드 링크 추출
                    p_tag = tr.find('p', attrs={'data-srvcid': True})
                    if p_tag:
                        data_attributes = p_tag.attrs
                        upper_no = data_attributes.get('data-upperno', '')
                        # 다운로드 링크 구성
                        download_link = f"https://seoulboard.seoul.go.kr/comm/getFile?srvcId=BBSTY1&upperNo={upper_no}&fileTy=ATTACH&fileNo=2&bbsNo=158"
                        download_links.append(download_link)

                        # 미리보기 링크 추출 (버튼의 data-url 속성 사용)
                        preview_button = tr.find(
                            'button', {'data-type': 'preview'})
                        if preview_button:
                            preview_link = preview_button.get('data-url', '')
                            print(preview_link)
                            preview_links.append(preview_link)

                # 첫 번째 다운로드 링크와 미리보기 링크 저장
                entry['download_link'] = download_links[0] if download_links else None
                entry['preview_link'] = preview_links[0] if preview_links else None

            except Exception as e:
                print(f"상세 페이지 {href} 크롤링 중 오류 발생: {e}")
                data_list[idx - 1]['download_link'] = None
                data_list[idx - 1]['preview_link'] = None
                continue

        # 상세 페이지 크롤링이 끝났으므로 브라우저 닫기
        await browser.close()

    # 3단계: 데이터 저장 (Excel 파일로) 및 열 너비 조정
    try:
        # 'data_code'와 'href'를 제외한 데이터만 추출
        data_without_href = [
            {
                'title': d['title'],
                'department': d['department'],
                'date': d['date'],
                'download_link': d['download_link'],
                'preview_link': d['preview_link']
            }
            for d in data_list
        ]

        # 데이터프레임 생성
        df = pd.DataFrame(data_without_href)

        # Excel 파일로 저장 (열 너비 자동 조정)
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # 각 열의 최대 길이를 계산하여 열 너비를 설정
            for i, column in enumerate(df.columns, 1):
                # 열의 최대 길이 계산
                max_length = df[column].astype(str).map(len).max()
                # 열 제목의 길이와 비교하여 더 큰 값 선택
                max_length = max(max_length, len(column)) + 2  # 여유 공간 추가
                # 열 번호를 열 문자로 변환
                column_letter = get_column_letter(i)
                # 열 너비 설정
                worksheet.column_dimensions[column_letter].width = max_length

        # 하이퍼링크 추가
        wb = load_workbook(excel_file)
        ws = wb['Sheet1']

        # 'download_link'와 'preview_link' 컬럼 인덱스 찾기
        download_col = df.columns.get_loc('download_link') + 1
        preview_col = df.columns.get_loc('preview_link') + 1

        # 하이퍼링크 설정
        for row in range(2, len(df) + 2):  # 헤더 제외
            # 다운로드 링크 하이퍼링크 설정
            download_cell = ws.cell(row=row, column=download_col)
            if download_cell.value and isinstance(download_cell.value, str) and download_cell.value.startswith('http'):
                download_cell.hyperlink = download_cell.value
                download_cell.font = Font(color='0000FF', underline='single')

            # 미리보기 링크 하이퍼링크 설정
            preview_cell = ws.cell(row=row, column=preview_col)
            if preview_cell.value and isinstance(preview_cell.value, str) and preview_cell.value.startswith('http'):
                preview_cell.hyperlink = preview_cell.value
                preview_cell.font = Font(color='0000FF', underline='single')

        # 워크북 저장
        wb.save(excel_file)

        print(f"\n총 {len(df)}개의 데이터가 '{excel_file}'에 저장되었습니다.")

    except Exception as e:
        print(f"엑셀 파일 저장 중 오류 발생: {e}")

if __name__ == "__main__":
    asyncio.run(main())




